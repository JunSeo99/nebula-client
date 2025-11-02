"""Spreadsheet extraction utilities.

This module provides helpers to read Excel or CSV files, derive lightweight
signals that can be fed to GPT prompts, and (optionally) enrich them with OCR
text pulled from embedded images or drawing text boxes.  The functions follow
common patterns used across the extraction package: errors raise a dedicated
exception type, optional dependencies are loaded lazily, and complex logic is
split into small, well-named helpers for easier testing.
"""

from __future__ import annotations

import io
import os
import re
import unicodedata
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

import numpy as np
import pandas as pd
from PIL import Image


class SpreadsheetExtractionError(RuntimeError):
    """Raised when we cannot analyse a spreadsheet."""


# ---------------------------------------------------------------------------
# Text normalisation helpers
# ---------------------------------------------------------------------------
_URL_PATTERN = re.compile(r"https?://|www\\.", re.IGNORECASE)
_UNNAMED_HEADER_PATTERN = re.compile(r"^Unnamed:? ?\\d+$", re.IGNORECASE)
_ASCII_NOISE_PATTERN = re.compile(r"^[A-Za-z\\s]{6,}$")
_HANGUL_RANGE = ("가", "힣")


def _nfkc(value: object) -> str:
    """Return text normalised with NFKC and stripped."""

    return unicodedata.normalize("NFKC", str(value or "")).strip()


def _tokenize(text: str) -> list[str]:
    """Tokenise text into alphanumeric or Hangul tokens."""

    if not text:
        return []

    normalised = _nfkc(text)
    if _URL_PATTERN.search(normalised):
        return []

    tokens = re.split(r"[^\\w가-힣]+", normalised)
    return [tok for tok in tokens if 2 <= len(tok) <= 30 and not tok.isdigit()]


def _split_words(text: str) -> list[str]:
    """Split camelCase/snake_case headers into human friendly words."""

    words = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", _nfkc(text))
    words = re.sub(r"[_\\-]+", " ", words)
    return words.split()


def _schema_strength(headers: Sequence[str]) -> float:
    """Estimate schema clarity based on the proportion of valid headers."""

    if not headers:
        return 0.0

    invalid = 0
    for header in headers:
        stripped = str(header).strip()
        if _UNNAMED_HEADER_PATTERN.match(stripped) or len(stripped) <= 1:
            invalid += 1
    return 1.0 - invalid / max(len(headers), 1)


def _hangul_ratio(text: str) -> float:
    """Return the ratio of Hangul characters in ``text``."""

    if not text:
        return 0.0
    normalised = _nfkc(text)
    count = sum(_HANGUL_RANGE[0] <= ch <= _HANGUL_RANGE[1] for ch in normalised)
    return count / max(1, len(normalised))


# ---------------------------------------------------------------------------
# File loading helpers
# ---------------------------------------------------------------------------

def load_first_sheet(path: str, *, max_rows: int = 20_000) -> pd.DataFrame:
    """Return a DataFrame built from the first worksheet or CSV content."""

    extension = Path(path).suffix.lower()

    try:
        if extension == ".csv":
            return _load_csv(path, max_rows)
        if extension in {".xlsx", ".xls", ".xlsm"}:
            return _load_excel(path, max_rows)
    except Exception as exc:  # pragma: no cover - surfaced to caller
        raise SpreadsheetExtractionError("스프레드시트 파일을 읽을 수 없습니다.") from exc

    raise SpreadsheetExtractionError("지원하지 않는 파일 형식입니다. csv/xlsx/xls/xlsm만 허용됩니다.")


def _load_csv(path: str, max_rows: int) -> pd.DataFrame:
    try:
        return pd.read_csv(path, nrows=max_rows)
    except Exception:
        pass

    for encoding in ("cp949", "utf-8-sig"):
        try:
            return pd.read_csv(path, nrows=max_rows, encoding=encoding, engine="python")
        except Exception:
            continue

    return pd.read_csv(path, nrows=max_rows, encoding_errors="ignore", engine="python")


def _load_excel(path: str, max_rows: int) -> pd.DataFrame:
    xl = pd.ExcelFile(path)
    return xl.parse(xl.sheet_names[0], nrows=max_rows)


# ---------------------------------------------------------------------------
# OCR helpers (best-effort, optional dependencies loaded lazily)
# ---------------------------------------------------------------------------

def _clean_ocr_text(text: str) -> str:
    cleaned = _nfkc(text)
    cleaned = re.sub(r"https?://\\S+|www\\.\\S+", " ", cleaned)
    cleaned = re.sub(r"[^0-9A-Za-z가-힣\\s]", " ", cleaned)
    return re.sub(r"\\s+", " ", cleaned).strip()


def _filter_ocr_lines(lines: Iterable[tuple[str, float]] | Iterable[str], *, dominant: str) -> list[str]:
    filtered: list[str] = []
    for line in lines:
        if isinstance(line, tuple):
            text, confidence = line[0], float(line[1])
            if confidence < 0.55:
                continue
        else:
            text = line
        cleaned = _clean_ocr_text(text)
        if not cleaned:
            continue
        if dominant == "hangul" and _hangul_ratio(cleaned) < 0.25 and _ASCII_NOISE_PATTERN.match(cleaned):
            continue
        if len(cleaned) <= 1:
            continue
        filtered.append(cleaned)

    unique: list[str] = []
    seen: set[str] = set()
    for candidate in filtered:
        lowered = candidate.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        unique.append(candidate)
    return unique


def _extract_text_boxes(path: str) -> list[str]:
    if not path.lower().endswith((".xlsx", ".xlsm")):
        return []

    texts: list[str] = []
    try:
        with zipfile.ZipFile(path, "r") as archive:
            for name in archive.namelist():
                if not name.startswith("xl/drawings/drawing") or not name.endswith(".xml"):
                    continue
                try:
                    data = archive.read(name)
                    import xml.etree.ElementTree as ET

                    root = ET.fromstring(data)
                except Exception:
                    continue
                for element in root.iter():
                    if element.tag.endswith("}t") and element.text:
                        cleaned = _clean_ocr_text(element.text)
                        if cleaned:
                            texts.append(cleaned)
    except Exception:
        return []

    unique: list[str] = []
    seen: set[str] = set()
    for candidate in texts:
        lowered = candidate.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        unique.append(candidate)
    return unique


def _extract_ocr_lines(path: str) -> list[tuple[str, float]]:
    """Extract OCR lines from embedded images inside an Excel workbook."""

    extension = Path(path).suffix.lower()
    if extension not in {".xls", ".xlsx", ".xlsm"}:
        return []

    try:
        import openpyxl  # type: ignore
    except Exception:
        return []

    try:
        import easyocr  # type: ignore
    except Exception:
        return []

    images: list[Image.Image] = []

    # Images referenced by openpyxl (drawing objects)
    try:
        workbook = openpyxl.load_workbook(path, data_only=True)
        sheet = workbook.active
        for drawing in getattr(sheet, "_images", []) or []:
            try:
                if hasattr(drawing, "_data") and callable(drawing._data):  # pragma: no cover - openpyxl internals
                    data = drawing._data()
                elif hasattr(drawing, "image") and hasattr(drawing.image, "blob"):
                    data = drawing.image.blob  # type: ignore[attr-defined]
                else:
                    continue
                images.append(Image.open(io.BytesIO(data)).convert("RGB"))
            except Exception:
                continue
    except Exception:
        pass

    # Images stored under xl/media
    try:
        with zipfile.ZipFile(path, "r") as archive:
            for name in archive.namelist():
                if not name.startswith("xl/media/"):
                    continue
                if not name.lower().endswith((".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tif", ".tiff")):
                    continue
                try:
                    data = archive.read(name)
                    images.append(Image.open(io.BytesIO(data)).convert("RGB"))
                except Exception:
                    continue
    except Exception:
        pass

    unique: list[Image.Image] = []
    seen: set[tuple[tuple[int, int], tuple[str, ...]]] = set()
    for image in images:
        key = (image.size, image.getbands())
        if key in seen or image.width < 40 or image.height < 20:
            continue
        seen.add(key)
        unique.append(image)

    if not unique:
        return []

    reader = easyocr.Reader(["ko", "en"], gpu=False)  # type: ignore
    results: list[tuple[str, float]] = []

    def _read(detail: int) -> None:
        for image in unique:
            for scale in (1.0, 1.5, 2.0):
                resized = (
                    image
                    if scale == 1.0
                    else image.resize((int(image.width * scale), int(image.height * scale)))
                )
                try:
                    if detail == 1:
                        for _bbox, text, score in reader.readtext(
                            np.array(resized),
                            detail=1,
                            paragraph=False,
                            text_threshold=0.7,
                            low_text=0.3,
                            mag_ratio=1.5,
                        ):
                            results.append((text, float(score)))
                    else:
                        for text in reader.readtext(
                            np.array(resized), detail=0, paragraph=False
                        ):
                            results.append((text, 0.99))
                except Exception:
                    continue

    _read(detail=1)
    if len(results) < 3:
        _read(detail=0)

    return results


# ---------------------------------------------------------------------------
# Feature extraction from spreadsheet content
# ---------------------------------------------------------------------------

def _dominant_script(strings: Iterable[str]) -> str:
    hangul = sum(any("가" <= ch <= "힣" for ch in token) for token in strings)
    latin = sum(any("A" <= ch <= "z" for ch in token) for token in strings)
    if hangul >= latin * 1.2 and hangul >= 5:
        return "hangul"
    if latin >= hangul * 1.2 and latin >= 5:
        return "latin"
    return "mixed"


def _collect_banner_tokens(path: str, *, use_ocr: bool) -> list[str]:
    extension = Path(path).suffix.lower()
    if extension not in {".xlsx", ".xls", ".xlsm"}:
        return []

    try:
        import openpyxl  # type: ignore
    except Exception:
        openpyxl = None  # type: ignore

    tokens: Counter[str] = Counter()

    if openpyxl:
        try:
            workbook = openpyxl.load_workbook(path, data_only=True)
            sheet = workbook.active
            merged_map: dict[tuple[int, int], tuple[int, int]] = {}
            try:
                merged_ranges = getattr(sheet.merged_cells, "ranges", sheet.merged_cells)
                for merged in merged_ranges:
                    for row in sheet.iter_rows(min_row=merged.min_row, max_row=merged.max_row,
                                               min_col=merged.min_col, max_col=merged.max_col):
                        for cell in row:
                            merged_map[(cell.row, cell.column)] = (merged.min_row, merged.min_col)
            except Exception:
                merged_map = {}

            max_rows = min(sheet.max_row or 0, 12)
            max_cols = min(sheet.max_column or 0, 8)

            for row_index in range(1, max_rows + 1):
                for col_index in range(1, max_cols + 1):
                    key = (row_index, col_index)
                    target = merged_map.get(key, key)
                    cell = sheet.cell(*target)
                    value = _nfkc(cell.value)
                    if not value or _URL_PATTERN.search(value):
                        continue

                    try:
                        font = cell.font
                        weight = 1.0 + 0.02 * (float(font.sz or 11.0) - 11.0)
                        if font.bold:
                            weight += 0.5
                        color_type = getattr(font.color, "type", None)
                        color_theme = getattr(font.color, "theme", None)
                        if color_type == "theme" and color_theme not in (None, 1):
                            weight += 0.3
                    except Exception:
                        weight = 1.0

                    for token in _tokenize(value):
                        tokens[token] += weight
        except Exception:
            tokens.clear()

    if use_ocr:
        ocr_lines = _filter_ocr_lines(_extract_ocr_lines(path), dominant="hangul")
        for ocr_token in _tokenize(" ".join(ocr_lines)):
            tokens[ocr_token] += 0.4

    return [token for token, _count in tokens.most_common(60)]


def _collect_sections(path: str) -> list[str]:
    extension = Path(path).suffix.lower()
    if extension not in {".xlsx", ".xls", ".xlsm"}:
        return []

    try:
        import openpyxl  # type: ignore
    except Exception:
        return []

    try:
        workbook = openpyxl.load_workbook(path, data_only=True)
        sheet = workbook.active
    except Exception:
        return []

    candidates: list[tuple[str, float]] = []
    max_rows = min(sheet.max_row or 0, 12)
    max_cols = min(sheet.max_column or 0, 8)

    for row_index in range(1, max_rows + 1):
        for col_index in range(1, max_cols + 1):
            cell = sheet.cell(row_index, col_index)
            value = _nfkc(cell.value)
            if not value or _URL_PATTERN.search(value):
                continue
            try:
                font = cell.font
                weight = 1.0 + 0.02 * (float(font.sz or 11.0) - 11.0)
                if font.bold:
                    weight += 0.6
            except Exception:
                weight = 1.0
            if re.search(r"[.!?]\\s*$", value) or len(value.split()) >= 10:
                weight *= 0.5
            candidates.append((value, weight))

    aggregator: dict[str, float] = defaultdict(float)
    for text, score in candidates:
        cleaned = re.sub(r"\\s+", " ", text).strip()
        if len(cleaned) <= 1 or _URL_PATTERN.search(cleaned):
            continue
        aggregator[cleaned] += score

    ranked = sorted(aggregator.items(), key=lambda item: (-item[1], -len(item[0])))
    results: list[str] = []
    seen: set[str] = set()
    for candidate, _score in ranked:
        lowered = candidate.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        results.append(candidate)
        if len(results) >= 6:
            break
    return results


def _sample_value_terms(frame: pd.DataFrame) -> list[str]:
    bag: Counter[str] = Counter()
    for column in frame.columns[:12]:
        series = frame[column].dropna().astype(str)
        series = series[series.str.len().between(2, 64)]
        series = series[~series.str.contains(_URL_PATTERN)]
        for value in series.head(500):
            for token in _tokenize(value):
                bag[token] += 1
    stop_words = {"및", "등", "자료", "현황", "총계", "합계", "확인", "공지", "입니다"}
    return [token for token, _count in bag.most_common(80) if token.lower() not in stop_words]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

@dataclass
class SpreadsheetSignals:
    """Describes extracted signals for GPT prompts."""

    title: str
    size: str
    headers: list[str]
    sections: list[str]
    banner: list[str]
    samples: list[str]


def build_summary_text(
    path: str,
    *,
    include_banner: bool = True,
    include_samples: bool = True,
    include_sections: bool = True,
    use_ocr: bool = True,
) -> tuple[str, SpreadsheetSignals]:
    """Return a condensed description of a spreadsheet and accompanying signals."""

    frame = load_first_sheet(path)
    rows, cols = frame.shape

    raw_headers = [str(column) for column in frame.columns]
    readable_headers = [
        " ".join(_split_words(header))
        for header in raw_headers
        if not _UNNAMED_HEADER_PATTERN.match(str(header))
    ]
    readable_headers = [header for header in readable_headers if header]
    headers_preview = readable_headers[:12]

    title = re.sub(
        r"\\s+",
        " ",
        os.path.splitext(os.path.basename(path))[0],
    ).strip()
    parts = [f"title:{title}" if title else None, f"size:{rows}x{cols}"]
    if headers_preview:
        parts.append("columns:" + ", ".join(headers_preview))

    banner_tokens: list[str] = []
    section_candidates: list[str] = []
    sample_terms: list[str] = []

    if include_banner:
        banner_tokens = _collect_banner_tokens(path, use_ocr=use_ocr)
    if include_sections:
        section_candidates = _collect_sections(path)
    if include_samples:
        sample_terms = _sample_value_terms(frame)

    dominant = _dominant_script(section_candidates + banner_tokens + headers_preview)
    text_box_lines = _extract_text_boxes(path) if use_ocr else []
    image_lines_raw = _extract_ocr_lines(path) if use_ocr else []
    image_lines = _filter_ocr_lines(image_lines_raw, dominant=dominant) if use_ocr else []

    if use_ocr:
        ocr_tokens: list[str] = []
        for line in text_box_lines + image_lines:
            for token in _tokenize(line):
                if _hangul_ratio(token) > 0.6 and len(token) >= 2:
                    ocr_tokens.append(token)
        unique_ocr_tokens: list[str] = []
        seen_tokens: set[str] = set()
        for token in ocr_tokens:
            lowered = token.lower()
            if lowered in seen_tokens:
                continue
            seen_tokens.add(lowered)
            unique_ocr_tokens.append(token)
            if len(unique_ocr_tokens) >= 6:
                break
        for token in reversed(unique_ocr_tokens):
            if token not in section_candidates:
                section_candidates.insert(0, token)
        banner_tokens = list(dict.fromkeys(banner_tokens + unique_ocr_tokens))

    if section_candidates:
        parts.append("sections:" + " ".join(section_candidates[:6]))
    if banner_tokens:
        parts.append("banner:" + " ".join(banner_tokens[:20]))

    if title and _schema_strength(raw_headers) < 0.5:
        parts = [f"title:{title}"] + [part for part in parts if part] + [f"title:{title}"]
    else:
        parts = [part for part in parts if part]

    description = " | ".join(parts)
    description = re.sub(r"\\s+", " ", description).strip()

    signals = SpreadsheetSignals(
        title=title,
        size=f"{rows}x{cols}",
        headers=headers_preview,
        sections=section_candidates,
        banner=banner_tokens,
        samples=sample_terms,
    )

    return description, signals


__all__ = [
    "SpreadsheetExtractionError",
    "SpreadsheetSignals",
    "build_summary_text",
]
